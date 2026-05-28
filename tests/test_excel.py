"""Unit tests for profiler.excel."""

from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from profiler.catalog import VolumeRef
from profiler.compare import compare_tables
from profiler.excel import write_workbook
from profiler.metamodel import (
    Alert,
    CategoricalStats,
    ColumnProfile,
    DatasetProfile,
    Lineage,
    NumericStats,
    ProfilerRun,
    new_run_id,
)
from profiler.storage import RunFolder


# ---------------------------------------------------------------------------
# Builders


def _numeric(**kw) -> NumericStats:
    base = dict(
        min=0.0, max=100.0, mean=50.0, stddev=20.0,
        p1=1.0, p5=5.0, p25=25.0, p50=50.0, p75=75.0, p95=95.0, p99=99.0,
        skewness=0.5, kurtosis=3.0,
        histogram_edges=[0.0, 50.0, 100.0],
        histogram_counts=[40, 60],
    )
    base.update(kw)
    return NumericStats(**base)


def _cat(**kw) -> CategoricalStats:
    base = dict(top_k={"A": 60, "B": 40}, entropy=1.0)
    base.update(kw)
    return CategoricalStats(**base)


def _col(name: str, logical_type: str = "string", nullable: bool = True, **kw) -> ColumnProfile:
    return ColumnProfile(
        name=name, logical_type=logical_type,
        physical_type=logical_type.upper(),
        nullable=nullable, null_count=0, null_pct=0.0,
        distinct_count=10, distinct_pct=0.1,
        **kw,
    )


def _dataset(cols: list[ColumnProfile], env: str = "PROD") -> DatasetProfile:
    return DatasetProfile(
        env_label=env, connection="local",
        catalog="dev", schema="prod_main_claims", table="medical_claim",
        row_count=50_000, column_count=len(cols), columns=cols,
    )


def _run(tmp_path: Path) -> tuple[ProfilerRun, RunFolder]:
    alert = Alert(rule="missing", severity="warn", message="5% null")
    cols_a = [
        _col("claim_id", nullable=False),
        _col("charge_amount", "double", numeric=_numeric(mean=1400.0)),
        _col("claim_type", categorical=_cat()),
        _col("old_col"),
    ]
    cols_b = [
        _col("claim_id", nullable=False),
        _col("charge_amount", "double", numeric=_numeric(mean=1900.0)),
        _col("claim_type", categorical=_cat(), alerts=[alert]),
        _col("new_col"),
    ]
    side_a = _dataset(cols_a, "PROD")
    side_b = DatasetProfile(
        env_label="TEST", connection="local",
        catalog="dev", schema="test_main_claims", table="medical_claim",
        row_count=50_000, column_count=len(cols_b), columns=cols_b,
    )
    comparisons = compare_tables(side_a, side_b)
    run = ProfilerRun(
        run_id=new_run_id(),
        created_utc=datetime.now(timezone.utc),
        side_a=side_a,
        side_b=side_b,
        comparisons=comparisons,
        lineage=Lineage(excel_summary="ab_summary.xlsx"),
    )
    folder_path = tmp_path / "run_001"
    folder_path.mkdir()
    folder = RunFolder(
        volume=VolumeRef(catalog="dev", schema="test_main_profiler", volume="ab_runs"),
        run_id="2025-01-01_1200",
        folder_name="2025-01-01_1200__prod-vs-test__medical_claim",
        path=str(folder_path),
    )
    return run, folder


# ---------------------------------------------------------------------------
# write_workbook — file creation


class TestWriteWorkbook:
    def test_creates_xlsx_file(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        assert Path(path).exists()
        assert path.endswith(".xlsx")

    def test_custom_filename(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run, filename="custom.xlsx")
        assert Path(path).name == "custom.xlsx"

    def test_readable_as_workbook(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        wb = openpyxl.load_workbook(path)
        assert wb is not None


# ---------------------------------------------------------------------------
# Sheet names


class TestSheetNames:
    def setup_method(self, method):
        pass  # tmp_path is pytest-injected per test

    def _wb(self, tmp_path) -> openpyxl.Workbook:
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)

    def test_has_overview_sheet(self, tmp_path):
        assert "Overview" in self._wb(tmp_path).sheetnames

    def test_has_schema_diff_sheet(self, tmp_path):
        assert "SchemaDiff" in self._wb(tmp_path).sheetnames

    def test_has_column_metrics_sheet(self, tmp_path):
        assert "ColumnMetrics" in self._wb(tmp_path).sheetnames

    def test_has_alerts_sheet(self, tmp_path):
        assert "Alerts" in self._wb(tmp_path).sheetnames

    def test_has_drift_scores_sheet(self, tmp_path):
        assert "DriftScores" in self._wb(tmp_path).sheetnames

    def test_five_sheets_total(self, tmp_path):
        assert len(self._wb(tmp_path).sheetnames) == 5


# ---------------------------------------------------------------------------
# Overview sheet


class TestOverviewSheet:
    def _ws(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)["Overview"]

    def _cell_values(self, ws) -> set[str]:
        return {str(ws.cell(r, c).value) for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1) if ws.cell(r, c).value is not None}

    def test_contains_run_id(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        ws = openpyxl.load_workbook(path)["Overview"]
        vals = self._cell_values(ws)
        assert str(run.run_id) in vals

    def test_contains_side_a_fqn(self, tmp_path):
        ws = self._ws(tmp_path)
        vals = self._cell_values(ws)
        assert "dev.prod_main_claims.medical_claim" in vals

    def test_contains_row_counts(self, tmp_path):
        ws = self._ws(tmp_path)
        vals = self._cell_values(ws)
        assert "50,000" in vals


# ---------------------------------------------------------------------------
# SchemaDiff sheet


class TestSchemaDiffSheet:
    def _ws(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)["SchemaDiff"]

    def _all_values(self, ws) -> list[str]:
        return [str(ws.cell(r, c).value)
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None]

    def test_header_row_present(self, tmp_path):
        ws = self._ws(tmp_path)
        row1 = [ws.cell(1, c).value for c in range(1, 8)]
        assert "Column" in row1

    def test_added_column_present(self, tmp_path):
        assert "new_col" in self._all_values(self._ws(tmp_path))

    def test_removed_column_present(self, tmp_path):
        assert "old_col" in self._all_values(self._ws(tmp_path))

    def test_unchanged_column_present(self, tmp_path):
        assert "claim_id" in self._all_values(self._ws(tmp_path))

    def test_row_count_matches_comparisons(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        ws = openpyxl.load_workbook(path)["SchemaDiff"]
        # header row + one data row per comparison
        assert ws.max_row == len(run.comparisons) + 1


# ---------------------------------------------------------------------------
# ColumnMetrics sheet


class TestColumnMetricsSheet:
    def _ws(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)["ColumnMetrics"]

    def _all_values(self, ws) -> list[str]:
        return [str(ws.cell(r, c).value)
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None]

    def test_charge_amount_column_present(self, tmp_path):
        assert "charge_amount" in self._all_values(self._ws(tmp_path))

    def test_numeric_mean_shown(self, tmp_path):
        vals = self._all_values(self._ws(tmp_path))
        # Side A mean ~1400, Side B mean ~1900
        assert any("1400" in v or "1,400" in v for v in vals)

    def test_header_includes_mean(self, tmp_path):
        ws = self._ws(tmp_path)
        row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert any("mean" in str(v).lower() for v in row1 if v)


# ---------------------------------------------------------------------------
# Alerts sheet


class TestAlertsSheet:
    def _ws(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)["Alerts"]

    def _all_values(self, ws) -> list[str]:
        return [str(ws.cell(r, c).value)
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if ws.cell(r, c).value is not None]

    def test_alert_rule_present(self, tmp_path):
        assert "missing" in self._all_values(self._ws(tmp_path))

    def test_side_label_present(self, tmp_path):
        vals = self._all_values(self._ws(tmp_path))
        assert "B" in vals  # alert was on side B

    def test_column_name_in_alert_row(self, tmp_path):
        assert "claim_type" in self._all_values(self._ws(tmp_path))


# ---------------------------------------------------------------------------
# DriftScores sheet (placeholder)


class TestDriftScoresSheet:
    def _ws(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        return openpyxl.load_workbook(path)["DriftScores"]

    def _all_text(self, ws) -> str:
        return " ".join(
            str(ws.cell(r, c).value)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value
        )

    def test_psi_header_present(self, tmp_path):
        assert "PSI" in self._all_text(self._ws(tmp_path))

    def test_verdict_header_present(self, tmp_path):
        assert "Verdict" in self._all_text(self._ws(tmp_path))

    def test_column_names_present(self, tmp_path):
        ws = self._ws(tmp_path)
        assert "claim_id" in self._all_text(ws)

    def test_row_count_matches_comparisons(self, tmp_path):
        run, folder = _run(tmp_path)
        path = write_workbook(folder, run)
        ws = openpyxl.load_workbook(path)["DriftScores"]
        assert ws.max_row == len(run.comparisons) + 1  # header + data rows
