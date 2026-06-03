"""Excel workbook writer (milestone 3).

Produces ab_summary.xlsx with five sheets:
    Overview      — run metadata, row/column counts, schema-change summary
    SchemaDiff    — one row per column: types, nullable flags, change verdict
    ColumnMetrics — side-by-side stats for columns present on both sides
    Alerts        — all alerts from both profiles
    DriftScores   — placeholder (PSI/KS/Chi-square/JS divergence in milestone 4)

All formatting is handled via openpyxl.  The workbook is written to the run
folder via storage.write_text (binary mode).
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .compare import schema_change_counts
from .metamodel import (
    ColumnComparison,
    ColumnProfile,
    DatasetProfile,
    ProfilerRun,
)
from .row_diff import RowDiffResult
from .storage import RunFolder


# ---------------------------------------------------------------------------
# Colour palette

_GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
_YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
_RED_FILL    = PatternFill("solid", fgColor="FFD7D7")
_GREEN_FILL  = PatternFill("solid", fgColor="D9EAD3")
_BLUE_FILL   = PatternFill("solid", fgColor="CFE2F3")

_HEADER_FONT  = Font(bold=True)
_TITLE_FONT   = Font(bold=True, size=13)

_VERDICT_FILL = {
    "stable":        _GREEN_FILL,
    "moderate":      _YELLOW_FILL,
    "significant":   _RED_FILL,
    "schema_change": _GREY_FILL,
}

_CHANGE_FILL = {
    "unchanged":          None,
    "added":              _GREEN_FILL,
    "removed":            _RED_FILL,
    "type_changed":       _YELLOW_FILL,
    "nullability_changed": _YELLOW_FILL,
}


# ---------------------------------------------------------------------------
# Public API


def write_workbook(
    folder: RunFolder,
    run: ProfilerRun,
    filename: str = "ab_summary.xlsx",
    row_diff: Optional[RowDiffResult] = None,
) -> str:
    """Write the Excel summary workbook. Returns the file path."""
    from .storage import write_bytes

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    _sheet_overview(wb, run)
    _sheet_schema_diff(wb, run)
    _sheet_column_metrics(wb, run)
    _sheet_alerts(wb, run)
    _sheet_drift_scores(wb, run)
    if row_diff is not None:
        _sheet_row_diff(wb, row_diff)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return write_bytes(folder, filename, buf.read())


# ---------------------------------------------------------------------------
# Sheet builders


def _sheet_overview(wb: openpyxl.Workbook, run: ProfilerRun) -> None:
    ws = wb.create_sheet("Overview")

    changes = schema_change_counts(run.comparisons)
    total_alerts_a = sum(len(c.alerts) for c in run.side_a.columns)
    total_alerts_b = sum(len(c.alerts) for c in run.side_b.columns)

    rows = [
        ("Run ID",                str(run.run_id)),
        ("Run label",             run.run_label or "—"),
        ("Created UTC",           run.created_utc.strftime("%Y-%m-%d %H:%M:%S UTC")),
        (None, None),
        ("Side A",                run.side_a.fqn),
        ("Side A env",            run.side_a.env_label),
        ("Side A row count",      f"{run.side_a.row_count:,}"),
        ("Side A column count",   run.side_a.column_count),
        ("Side A duplicate rows", f"{run.side_a.duplicate_rows:,}"),
        ("Side A alerts",         total_alerts_a),
        (None, None),
        ("Side B",                run.side_b.fqn),
        ("Side B env",            run.side_b.env_label),
        ("Side B row count",      f"{run.side_b.row_count:,}"),
        ("Side B column count",   run.side_b.column_count),
        ("Side B duplicate rows", f"{run.side_b.duplicate_rows:,}"),
        ("Side B alerts",         total_alerts_b),
        (None, None),
        ("Schema — unchanged",            changes["unchanged"]),
        ("Schema — added (B only)",       changes["added"]),
        ("Schema — removed (A only)",     changes["removed"]),
        ("Schema — type changed",         changes["type_changed"]),
        ("Schema — nullability changed",  changes["nullability_changed"]),
    ]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50

    title = ws.cell(row=1, column=1, value="A/B Profile Run — Overview")
    title.font = _TITLE_FONT
    ws.append([])

    for label, value in rows:
        if label is None:
            ws.append([])
        else:
            row_idx = ws.max_row + 1
            ws.append([label, value])
            ws.cell(row=row_idx, column=1).font = _HEADER_FONT


def _sheet_schema_diff(wb: openpyxl.Workbook, run: ProfilerRun) -> None:
    ws = wb.create_sheet("SchemaDiff")

    headers = [
        "Column", "Side A Type", "Side A Nullable",
        "Side B Type", "Side B Nullable",
        "Change", "Verdict",
    ]
    _write_header_row(ws, headers)

    a_map = {c.name: c for c in run.side_a.columns}
    b_map = {c.name: c for c in run.side_b.columns}

    for cmp in run.comparisons:
        col_a = a_map.get(cmp.column_name)
        col_b = b_map.get(cmp.column_name)
        row = [
            cmp.column_name,
            col_a.logical_type if col_a else "—",
            str(col_a.nullable).lower() if col_a else "—",
            col_b.logical_type if col_b else "—",
            str(col_b.nullable).lower() if col_b else "—",
            cmp.schema_change,
            cmp.verdict,
        ]
        ws.append(row)
        fill = _CHANGE_FILL.get(cmp.schema_change)
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

    _autofit(ws, headers)


def _sheet_column_metrics(wb: openpyxl.Workbook, run: ProfilerRun) -> None:
    ws = wb.create_sheet("ColumnMetrics")

    headers = [
        "Column", "Type",
        "A null%", "B null%", "Δ null%",
        "A distinct%", "B distinct%",
        "A mean", "B mean", "Δ mean%",
        "A stddev", "B stddev",
        "A p50", "B p50",
        "A entropy", "B entropy",
        "A top value", "B top value",
    ]
    _write_header_row(ws, headers)

    a_map = {c.name: c for c in run.side_a.columns}
    b_map = {c.name: c for c in run.side_b.columns}

    for cmp in run.comparisons:
        col_a = a_map.get(cmp.column_name)
        col_b = b_map.get(cmp.column_name)

        if col_a is None or col_b is None:
            # Schema-change column — no side-by-side stats
            ws.append([
                cmp.column_name,
                "—" if col_a is None else col_a.logical_type,
                *["—"] * (len(headers) - 2),
            ])
            fill = _CHANGE_FILL.get(cmp.schema_change)
            if fill:
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=ci).fill = fill
            continue

        a_null  = _pct(col_a.null_pct)
        b_null  = _pct(col_b.null_pct)
        d_null  = _pct(col_b.null_pct - col_a.null_pct, signed=True)

        a_dist  = _pct(col_a.distinct_pct)
        b_dist  = _pct(col_b.distinct_pct)

        a_mean = _fmt_num(col_a.numeric.mean if col_a.numeric else None)
        b_mean = _fmt_num(col_b.numeric.mean if col_b.numeric else None)
        d_mean = _mean_delta_pct(
            col_a.numeric.mean if col_a.numeric else None,
            col_b.numeric.mean if col_b.numeric else None,
        )
        a_std  = _fmt_num(col_a.numeric.stddev if col_a.numeric else None)
        b_std  = _fmt_num(col_b.numeric.stddev if col_b.numeric else None)
        a_p50  = _fmt_num(col_a.numeric.p50 if col_a.numeric else None)
        b_p50  = _fmt_num(col_b.numeric.p50 if col_b.numeric else None)

        a_ent  = _fmt_num(col_a.categorical.entropy if col_a.categorical else None, dp=3)
        b_ent  = _fmt_num(col_b.categorical.entropy if col_b.categorical else None, dp=3)

        a_top  = _top_value(col_a)
        b_top  = _top_value(col_b)

        ws.append([
            cmp.column_name, col_a.logical_type,
            a_null, b_null, d_null,
            a_dist, b_dist,
            a_mean, b_mean, d_mean,
            a_std,  b_std,
            a_p50,  b_p50,
            a_ent,  b_ent,
            a_top,  b_top,
        ])

        # Highlight rows where null% delta > 10 pp
        if col_a.null_pct is not None and col_b.null_pct is not None:
            if abs(col_b.null_pct - col_a.null_pct) > 0.10:
                for ci in (3, 4, 5):
                    ws.cell(row=ws.max_row, column=ci).fill = _YELLOW_FILL

    _autofit(ws, headers)


def _sheet_alerts(wb: openpyxl.Workbook, run: ProfilerRun) -> None:
    ws = wb.create_sheet("Alerts")

    headers = ["Side", "Column", "Rule", "Severity", "Message"]
    _write_header_row(ws, headers)

    sev_fill = {"critical": _RED_FILL, "warn": _YELLOW_FILL, "info": None}

    for side_label, profile in (("A", run.side_a), ("B", run.side_b)):
        for col in profile.columns:
            for alert in col.alerts:
                ws.append([side_label, col.name, alert.rule, alert.severity, alert.message])
                fill = sev_fill.get(alert.severity)
                if fill:
                    for ci in range(1, len(headers) + 1):
                        ws.cell(row=ws.max_row, column=ci).fill = fill

    _autofit(ws, headers)


def _sheet_drift_scores(wb: openpyxl.Workbook, run: ProfilerRun) -> None:
    ws = wb.create_sheet("DriftScores")

    headers = [
        "Column", "Type", "Schema change",
        "PSI", "Verdict",
        "KS stat", "KS p-value",
        "Chi-square", "JS divergence",
    ]
    _write_header_row(ws, headers)

    a_map = {c.name: c for c in run.side_a.columns}

    for cmp in run.comparisons:
        col_a = a_map.get(cmp.column_name)
        col_type = col_a.logical_type if col_a else "—"

        row = [
            cmp.column_name,
            col_type,
            cmp.schema_change,
            _fmt_num(cmp.psi, dp=4) if cmp.psi is not None else "—",
            cmp.verdict,
            _fmt_num(cmp.ks_stat, dp=4) if cmp.ks_stat is not None else "—",
            _fmt_num(cmp.ks_pvalue, dp=4) if cmp.ks_pvalue is not None else "—",
            _fmt_num(cmp.chi_square, dp=2) if cmp.chi_square is not None else "—",
            _fmt_num(cmp.js_divergence, dp=4) if cmp.js_divergence is not None else "—",
        ]
        ws.append(row)

        fill = _VERDICT_FILL.get(cmp.verdict)
        if fill:
            # Colour the entire row for drifted/schema-changed columns
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=ci).fill = fill

    _autofit(ws, headers)


# ---------------------------------------------------------------------------
# Formatting helpers


def _write_header_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    start_row: Optional[int] = None,
) -> None:
    if start_row is not None:
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _BLUE_FILL
    else:
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col_idx).font = _HEADER_FONT
            ws.cell(row=ws.max_row, column=col_idx).fill = _BLUE_FILL


def _autofit(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: list[str],
    min_width: int = 10,
    max_width: int = 40,
) -> None:
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = max(min_width, min(len(header) + 4, max_width))
        ws.column_dimensions[col_letter].width = width


def _pct(value: Optional[float], signed: bool = False) -> str:
    if value is None:
        return "—"
    prefix = "+" if signed and value > 0 else ""
    return f"{prefix}{value * 100:.1f}%"


def _fmt_num(value: Optional[float], dp: int = 2) -> str:
    if value is None:
        return "—"
    if abs(value) >= 1_000_000:
        return f"{value:,.0f}"
    if abs(value) >= 1_000:
        return f"{value:,.1f}"
    return f"{value:.{dp}f}"


def _mean_delta_pct(a: Optional[float], b: Optional[float]) -> str:
    if a is None or b is None:
        return "—"
    if a == 0:
        return "—"
    delta = (b - a) / abs(a) * 100
    prefix = "+" if delta > 0 else ""
    return f"{prefix}{delta:.1f}%"


def _top_value(col: ColumnProfile) -> str:
    if col.categorical and col.categorical.top_k:
        key = next(iter(col.categorical.top_k))
        count = col.categorical.top_k[key]
        return f"{key} ({count:,})"
    return "—"


def _sheet_row_diff(wb: openpyxl.Workbook, rd: RowDiffResult) -> None:
    ws = wb.create_sheet("RowDiff")

    # Summary section
    title = ws.cell(row=1, column=1, value="Row-Level Diff Summary")
    title.font = _TITLE_FONT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    summary_rows = [
        ("Key columns", ", ".join(rd.key_columns)),
        ("Rows removed (only in Side A)", f"{rd.rows_only_in_a:,}"),
        ("Rows added (only in Side B)",   f"{rd.rows_only_in_b:,}"),
        ("Rows changed (same key, diff values)", f"{rd.rows_changed:,}"),
        ("Rows identical",               f"{rd.rows_identical:,}"),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT

    if rd.error:
        ws.append([])
        ws.cell(row=ws.max_row + 1, column=1,
                 value=f"Error: {rd.error}").font = Font(color="FF0000")
        return

    col_names = rd.col_names or []

    # Removed rows sample
    if rd.sample_removed:
        ws.append([])
        ws.append(["--- Sample: Removed rows (in Side A, not in Side B) ---"])
        ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
        _write_header_row(ws, col_names)
        for row in rd.sample_removed:
            ws.append([str(v) if v is not None else "" for v in row])
            for ci in range(1, len(col_names) + 1):
                ws.cell(row=ws.max_row, column=ci).fill = _RED_FILL

    # Added rows sample
    if rd.sample_added:
        ws.append([])
        ws.append(["--- Sample: Added rows (in Side B, not in Side A) ---"])
        ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
        _write_header_row(ws, col_names)
        for row in rd.sample_added:
            ws.append([str(v) if v is not None else "" for v in row])
            for ci in range(1, len(col_names) + 1):
                ws.cell(row=ws.max_row, column=ci).fill = _GREEN_FILL

    # Changed rows sample (interleaved A/B)
    if rd.sample_changed:
        ws.append([])
        ws.append(["--- Sample: Changed rows (A row then B row per key) ---"])
        ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
        changed_headers = (rd.col_names or [])
        _write_header_row(ws, changed_headers)
        prev_key = None
        fill_toggle = True
        for row in rd.sample_changed:
            # Use first key column to detect pair boundaries
            key_val = row[0] if row else None
            if key_val != prev_key:
                fill_toggle = not fill_toggle
                prev_key = key_val
            fill = _BLUE_FILL if fill_toggle else _YELLOW_FILL
            ws.append([str(v) if v is not None else "" for v in row])
            for ci in range(1, len(changed_headers) + 1):
                ws.cell(row=ws.max_row, column=ci).fill = fill

    _autofit(ws, col_names or ["A"])


